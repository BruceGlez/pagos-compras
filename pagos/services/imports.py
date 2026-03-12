from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any
import unicodedata
import re

import xlrd
from openpyxl import load_workbook

from pagos.models import (
    Anticipo,
    Compra,
    ImportRowLog,
    ImportRun,
    MonedaChoices,
    PersonaFactura,
    Productor,
    SiNoChoices,
    WorkflowStateChoices,
)


def _norm_col(value: str) -> str:
    return " ".join((value or "").strip().upper().replace("_", " ").replace("\n", " ").split())


def _normalize_name(value: str) -> str:
    s = (value or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return " ".join(s.split())


def _name_signature(value: str) -> str:
    tokens = [t for t in _normalize_name(value).split(" ") if t]
    tokens.sort()
    return "|".join(tokens)


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    return Decimal(s or "0")


def _to_date(v: Any) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date fallback (works for old .xls reports)
        base = date(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            pass
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return date.today()


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    duplicates: int = 0
    divisions_created: int = 0
    error_count: int = 0
    conflict_count: int = 0


def _read_rows(path: str | Path):
    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb["COMPRAS"] if "COMPRAS" in wb.sheetnames else wb.active
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    except Exception:
        book = xlrd.open_workbook(path)
        sh = book.sheet_by_name("COMPRAS") if "COMPRAS" in book.sheet_names() else book.sheet_by_index(0)
        return [tuple(sh.row_values(i)) for i in range(sh.nrows)]


def _find_header_row(rows: list[tuple]) -> int:
    for i, row in enumerate(rows[:40]):
        cols = {_norm_col(str(c)) for c in row if str(c).strip()}
        if "COMPRA" in cols and "PRODUCTOR" in cols:
            return i
    return 0


def _build_parsed_records(path: str | Path):
    rows = _read_rows(path)
    if not rows:
        return [], {}

    h = _find_header_row(rows)
    headers = [_norm_col(str(c or "")) for c in rows[h]]
    idx = {h: i for i, h in enumerate(headers)}

    aliases = {
        "FECHA LIQ": ["FECHA LIQ", "FECHA"],
        # Prioridad: usar columna de total real de compra en dólares/libras cuando exista.
        "COMPRA EN LIBRAS": [
            "TOTAL DLLS LIBRAS",
            "TOTAL DLS LIBRAS",
            "TOTAL DOLLARS LIBRAS",
            "COMPRA EN LIBRAS",
            "TOTAL EN DLS",
            "TOTAL DLLS",
            "TOTAL DLS",
        ],
        "PACAS": ["PACAS", "CANT"],
        "RETENCION (DEUDAS) USD": ["RETENCION (DEUDAS) USD", "RETENCION", "RETENCIÓN"],
    }

    def val(r, key, default=None):
        keys = aliases.get(key, [key])
        for k in keys:
            i = idx.get(_norm_col(k))
            if i is None or i >= len(r):
                continue
            v = r[i]

            # Caso común en reportes Crystal: encabezado en la última columna,
            # pero el valor queda una celda a la izquierda por combinación de celdas.
            if key == "COMPRA EN LIBRAS" and (v is None or (isinstance(v, str) and not v.strip())):
                if i - 1 >= 0:
                    left = r[i - 1]
                    if left is not None and (not isinstance(left, str) or left.strip()):
                        return left

            # Para aliases, ignora celdas vacías y sigue buscando fallback.
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return v
        return default

    groups: dict[tuple, int] = {}
    parsed: list[tuple[int, tuple, dict]] = []

    for offset, row in enumerate(rows[h + 1 :], start=h + 2):
        numero = int(_to_decimal(val(row, "COMPRA", 0)))
        if numero <= 0:
            continue
        productor_nombre = str(val(row, "PRODUCTOR", "") or "").strip()
        if not productor_nombre:
            continue

        fecha_liq = _to_date(val(row, "FECHA LIQ", date.today()))
        compra_total = _to_decimal(val(row, "COMPRA EN LIBRAS", 0))

        rec = {
            "numero_compra": numero,
            "productor_nombre": productor_nombre,
            "fecha_liq": fecha_liq,
            "fecha_de_pago": _to_date(val(row, "FECHA DE PAGO", fecha_liq)),
            "regimen_fiscal": str(val(row, "REGIMEN FISCAL", "") or "").strip(),
            "factura": str(val(row, "FACTURA", "") or "").strip(),
            "uuid_factura": str(val(row, "UUID FACTURA", "") or "").strip(),
            "pacas": _to_decimal(val(row, "PACAS", 0)),
            "compra_en_libras": compra_total,
            "anticipo": _to_decimal(val(row, "ANTICIPO", 0)),
            "pago": _to_decimal(val(row, "PAGO", 0)),
            "retencion_deudas_usd": _to_decimal(val(row, "RETENCION (DEUDAS) USD", 0)),
            "retencion_deudas_mxn": _to_decimal(val(row, "RETENCION (DEUDAS) MXN", 0)),
            "retencion_resico": _to_decimal(val(row, "RETENCION RESICO 1.25%", 0)),
            "saldo_pendiente": _to_decimal(val(row, "SALDO PENDIENTE", 0)),
            "cuenta_de_pago": str(val(row, "CUENTA DE PAGO", "") or "").strip(),
            "metodo_de_pago": str(val(row, "METODO DE PAGO", "") or "").strip(),
            "cuenta_productor": str(val(row, "CUENTA PRODUCTOR", "") or "").strip(),
            "workflow_state": WorkflowStateChoices.IMPORTED,
            "intereses": SiNoChoices.SI
            if str(val(row, "INTERESES", "NO") or "NO").strip().upper() in {"SI", "S", "YES", "Y"}
            else SiNoChoices.NO,
        }

        key = (numero, _name_signature(productor_nombre), fecha_liq)
        groups[key] = groups.get(key, 0) + 1
        parsed.append((offset, key, rec))

    return parsed, groups


def preview_compras_excel(path: str | Path, *, limit: int = 20):
    parsed, groups = _build_parsed_records(path)
    out = []
    for row_number, key, rec in parsed[:limit]:
        out.append(
            {
                "row_number": row_number,
                "compra": rec["numero_compra"],
                "productor": rec["productor_nombre"],
                "fecha_liq": rec["fecha_liq"],
                "pacas": rec["pacas"],
                "total_dls": rec["compra_en_libras"],
                "es_division_detectada": groups.get(key, 0) > 1,
            }
        )
    return out


def _resolve_or_create_productor(nombre: str):
    n = _normalize_name(nombre)
    if not n:
        return None

    direct = Productor.objects.filter(nombre__iexact=nombre).first()
    if direct:
        return direct

    sig = _name_signature(nombre)
    for p in Productor.objects.all().only("id", "nombre"):
        if _name_signature(p.nombre) == sig:
            return p

    p = Productor(nombre=nombre, codigo="", activo=True)
    p.save()
    return p


def _resolve_or_create_persona(nombre: str):
    n = _normalize_name(nombre)
    if not n:
        return None

    direct = PersonaFactura.objects.filter(nombre__iexact=nombre).first()
    if direct:
        return direct

    sig = _name_signature(nombre)
    for p in PersonaFactura.objects.all().only("id", "nombre"):
        if _name_signature(p.nombre) == sig:
            return p

    return PersonaFactura.objects.create(nombre=nombre)


def detect_compras_conflicts(path: str | Path):
    parsed, _groups = _build_parsed_records(path)
    out = []
    for row_number, _key, rec0 in parsed:
        rec = dict(rec0)
        productor = _resolve_or_create_productor(rec.pop("productor_nombre"))
        existing = Compra.objects.filter(
            numero_compra=rec["numero_compra"],
            productor=productor,
            parent_compra__isnull=True,
        ).first()
        if not existing:
            continue
        same_payload = (
            (existing.fecha_liq == rec.get("fecha_liq"))
            and ((existing.pacas or Decimal("0")) == (rec.get("pacas") or Decimal("0")))
            and ((existing.compra_en_libras or Decimal("0")) == (rec.get("compra_en_libras") or Decimal("0")))
        )
        if same_payload:
            continue
        out.append(
            {
                "row_number": row_number,
                "numero_compra": rec.get("numero_compra"),
                "productor": productor.nombre,
                "existing_fecha": existing.fecha_liq,
                "incoming_fecha": rec.get("fecha_liq"),
                "existing_pacas": existing.pacas,
                "incoming_pacas": rec.get("pacas"),
                "existing_total": existing.compra_en_libras,
                "incoming_total": rec.get("compra_en_libras"),
            }
        )
    return out


def import_compras_excel(path: str | Path, *, dry_run: bool = False, conflict_policy: str = "ask", conflict_resolutions: dict | None = None) -> ImportStats:
    parsed, groups = _build_parsed_records(path)

    stats = ImportStats()
    base_by_key: dict[tuple, Compra] = {}
    run = ImportRun.objects.create(source_name=str(path), dry_run=dry_run)

    for row_number, key, rec0 in parsed:
        rec = dict(rec0)
        try:
            productor = _resolve_or_create_productor(rec.pop("productor_nombre"))
            rec["productor"] = productor

            existing = Compra.objects.filter(
                numero_compra=rec["numero_compra"],
                productor=productor,
                parent_compra__isnull=True,
            ).first()

            if key not in base_by_key:
                if existing:
                    # Same key imported again: detect exact duplicate vs conflict
                    same_payload = (
                        (existing.fecha_liq == rec.get("fecha_liq"))
                        and ((existing.pacas or Decimal("0")) == (rec.get("pacas") or Decimal("0")))
                        and ((existing.compra_en_libras or Decimal("0")) == (rec.get("compra_en_libras") or Decimal("0")))
                    )
                    if same_payload:
                        stats.duplicates += 1
                        base = existing
                        ImportRowLog.objects.create(
                            run=run,
                            row_number=row_number,
                            status="duplicate",
                            message="Compra idéntica ya existente: se omite.",
                            compra_numero=rec.get("numero_compra"),
                            productor_nombre=productor.nombre,
                        )
                    else:
                        stats.conflict_count += 1
                        row_policy = (conflict_resolutions or {}).get(str(row_number), conflict_policy)
                        if row_policy == "overwrite":
                            existing.fecha_liq = rec.get("fecha_liq")
                            existing.pacas = rec.get("pacas")
                            existing.compra_en_libras = rec.get("compra_en_libras")
                            existing.factura = rec.get("factura", "")
                            existing.uuid_factura = rec.get("uuid_factura", "")
                            if not dry_run:
                                existing.save()
                            stats.updated += 1
                            base = existing
                            ImportRowLog.objects.create(
                                run=run,
                                row_number=row_number,
                                status="updated",
                                message="Conflicto resuelto: se sobrescribió compra existente.",
                                compra_numero=rec.get("numero_compra"),
                                productor_nombre=productor.nombre,
                            )
                        elif row_policy == "keep_existing":
                            base = existing
                            ImportRowLog.objects.create(
                                run=run,
                                row_number=row_number,
                                status="conflict",
                                message="Conflicto detectado: se conservó compra existente.",
                                compra_numero=rec.get("numero_compra"),
                                productor_nombre=productor.nombre,
                            )
                        else:
                            base = existing
                            ImportRowLog.objects.create(
                                run=run,
                                row_number=row_number,
                                status="conflict",
                                message="Conflicto detectado: confirma política (conservar/sobrescribir).",
                                compra_numero=rec.get("numero_compra"),
                                productor_nombre=productor.nombre,
                            )
                else:
                    base = Compra(**rec)
                    if not dry_run:
                        base.save()
                    stats.created += 1
                    ImportRowLog.objects.create(
                        run=run,
                        row_number=row_number,
                        status="created",
                        message="Compra base creada",
                        compra_numero=rec.get("numero_compra"),
                        productor_nombre=productor.nombre,
                    )
                base_by_key[key] = base
                continue

            if groups.get(key, 0) <= 1:
                continue

            base = base_by_key[key]
            base_total = base.compra_en_libras or Decimal("0")
            pct = (rec["compra_en_libras"] * Decimal("100") / base_total) if base_total > 0 else Decimal("0")

            division = Compra(**rec, parent_compra=base, porcentaje_division=pct, workflow_state=WorkflowStateChoices.IMPORTED)
            if not dry_run:
                division.save()
            stats.divisions_created += 1
            stats.created += 1
            ImportRowLog.objects.create(
                run=run,
                row_number=row_number,
                status="division",
                message="División creada",
                compra_numero=rec.get("numero_compra"),
                productor_nombre=productor.nombre,
            )
        except Exception as e:
            stats.error_count = getattr(stats, "error_count", 0) + 1
            ImportRowLog.objects.create(
                run=run,
                row_number=row_number,
                status="error",
                message=str(e),
                compra_numero=rec.get("numero_compra"),
                productor_nombre=str(rec0.get("productor_nombre", "")),
            )

    run.created_count = stats.created
    run.duplicate_count = stats.duplicates
    run.division_count = stats.divisions_created
    run.error_count = getattr(stats, "error_count", 0)
    run.save(update_fields=["created_count", "duplicate_count", "division_count", "error_count", "updated_at"])

    return stats


def preview_anticipos_excel(path: str | Path, *, limit: int = 20):
    rows = _read_rows(path)
    if not rows:
        return []
    h = _find_header_row(rows)
    headers = [_norm_col(str(c or "")) for c in rows[h]]
    idx = {h: i for i, h in enumerate(headers)}

    aliases = {
        "ANTICIPO_NUM": ["ANTICIPO", "NO ANTICIPO", "NUMERO ANTICIPO"],
        "FECHA": ["FECHA DE PAGO", "FECHA"],
        "PRODUCTOR": ["PRODUCTOR"],
        "PERSONA": ["PERSONA QUE FACTURA"],
        "FACTURA": ["FACTURA"],
        "UUID_NC": ["UUID NOTA DE CREDITO", "UUID NOTA DE CRÉDITO", "UUID NC"],
        "MONTO": ["ANTICIPO", "MONTO ANTICIPO", "MONTO"],
        "MONEDA": ["MONEDA"],
    }

    def val(r, key, default=None):
        for k in aliases.get(key, [key]):
            i = idx.get(_norm_col(k))
            if i is not None and i < len(r):
                return r[i]
        return default

    out = []
    for rn, row in enumerate(rows[h + 1 :], start=h + 2):
        productor = str(val(row, "PRODUCTOR", "") or "").strip()
        monto = _to_decimal(val(row, "MONTO", 0))
        if not productor or monto <= 0:
            continue
        out.append(
            {
                "row_number": rn,
                "numero_anticipo": int(_to_decimal(val(row, "ANTICIPO_NUM", 0))) or None,
                "fecha_pago": _to_date(val(row, "FECHA", date.today())),
                "productor": productor,
                "persona_que_factura": str(val(row, "PERSONA", "") or "").strip(),
                "factura": str(val(row, "FACTURA", "") or "").strip(),
                "uuid_nota_credito": str(val(row, "UUID_NC", "") or "").strip(),
                "monto": monto,
                "moneda": str(val(row, "MONEDA", "DOLARES") or "DOLARES"),
            }
        )
        if len(out) >= limit:
            break
    return out


def import_anticipos_excel(path: str | Path, *, dry_run: bool = False) -> ImportStats:
    rows = _read_rows(path)
    if not rows:
        return ImportStats()
    h = _find_header_row(rows)
    headers = [_norm_col(str(c or "")) for c in rows[h]]
    idx = {h: i for i, h in enumerate(headers)}

    aliases = {
        "ANTICIPO_NUM": ["ANTICIPO", "NO ANTICIPO", "NUMERO ANTICIPO"],
        "FECHA": ["FECHA DE PAGO", "FECHA"],
        "PRODUCTOR": ["PRODUCTOR"],
        "PERSONA": ["PERSONA QUE FACTURA"],
        "FACTURA": ["FACTURA"],
        "MONTO": ["ANTICIPO", "MONTO ANTICIPO", "MONTO"],
        "MONEDA": ["MONEDA"],
    }

    def val(r, key, default=None):
        for k in aliases.get(key, [key]):
            i = idx.get(_norm_col(k))
            if i is not None and i < len(r):
                return r[i]
        return default

    stats = ImportStats()
    run = ImportRun.objects.create(source_name=f"{path}::ANTICIPOS", dry_run=dry_run)

    for rn, row in enumerate(rows[h + 1 :], start=h + 2):
        try:
            productor_nombre = str(val(row, "PRODUCTOR", "") or "").strip()
            monto = _to_decimal(val(row, "MONTO", 0))
            if not productor_nombre or monto <= 0:
                continue

            productor = _resolve_or_create_productor(productor_nombre)

            numero = int(_to_decimal(val(row, "ANTICIPO_NUM", 0)))
            fecha_pago = _to_date(val(row, "FECHA", date.today()))
            moneda_raw = str(val(row, "MONEDA", "DOLARES") or "DOLARES").upper()
            moneda = MonedaChoices.DOLARES if "DOL" in moneda_raw else MonedaChoices.PESOS

            existing = None
            if numero > 0:
                existing = Anticipo.objects.filter(numero_anticipo=numero).first()

            if existing:
                stats.duplicates += 1
                ImportRowLog.objects.create(run=run, row_number=rn, status="duplicate", message="Anticipo ya existe", compra_numero=numero, productor_nombre=productor_nombre)
                continue

            persona_txt = str(val(row, "PERSONA", "") or "").strip()
            persona_obj = _resolve_or_create_persona(persona_txt)

            ant = Anticipo(
                numero_anticipo=(numero if numero > 0 else None),
                fecha_pago=fecha_pago,
                productor=productor,
                persona_facturadora=persona_obj,
                persona_que_factura=persona_txt,
                factura=str(val(row, "FACTURA", "") or "").strip(),
                uuid_nota_credito=str(val(row, "UUID_NC", "") or "").strip(),
                monto_anticipo=monto,
                moneda=moneda,
            )
            if not dry_run:
                ant.save()
            stats.created += 1
            ImportRowLog.objects.create(run=run, row_number=rn, status="created", message="Anticipo creado", compra_numero=numero or None, productor_nombre=productor_nombre)
        except Exception as e:
            stats.error_count += 1
            ImportRowLog.objects.create(run=run, row_number=rn, status="error", message=str(e), productor_nombre=str(val(row, "PRODUCTOR", "") or ""))

    run.created_count = stats.created
    run.duplicate_count = stats.duplicates
    run.division_count = 0
    run.error_count = stats.error_count
    run.save(update_fields=["created_count", "duplicate_count", "division_count", "error_count", "updated_at"])
    return stats
